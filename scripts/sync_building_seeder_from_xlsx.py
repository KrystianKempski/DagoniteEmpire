#!/usr/bin/env python3
"""Rebuild BuildingTemplateSeeder.cs Add(...) list from Buildings sheet in Katalog_towarow_handlowych.xlsx."""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs/barony/Katalog_towarow_handlowych.xlsx"
SEEDER = ROOT / "DagoniteEmpire/Service/BuildingTemplateSeeder.cs"

PPB_COLS = [
    ("Food", "food"),
    ("Economy", "economy"),
    ("Production", "production"),
    ("Loyalty", "loyalty"),
    ("Stability", "stability"),
    ("Law", "law"),
    ("Corruption", "corruption"),
    ("Science", "science"),
    ("Magic", "magic"),
    ("Culture", "culture"),
    ("Intelligence", "intelligence"),
    ("Defense", "defense"),
    ("Treasury", "treasury"),
]

TERRAIN_BY_NAME = {
    "Quarry - Granite": "Quarry",
    "Quarry - common stone": "Quarry",
    "Quarry - Obsidian": "Obsidian",
    "Quarry - Tarnit": "Quarry",
    "Clay pit": "Clay",
    "Mine - precious gems (luxury)": "Quarry",
    "Mine - soft metals": "Mine",
    "Mine - Silver": "Mine",
    "Mine - Salt": "Salt",
    "Mine - Sulfur": "Sulfur",
    "Mine - Gold": "Mine",
    "Mine - Iron": "Mine",
    "Mine - Dagoferryt": "Mine",
    "Sawmill - common": "Forest",
    "Sawmill - Ironwood": "Forest",
    "Sawmill - Elven alder": "Forest",
    "Sawmill - Shipbuilding wood": "Forest",
}


def pn(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return float(str(v).strip().replace(",", "."))


def esc(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"')


def load_excel_buildings() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Buildings"]
    hdr = {ws.cell(1, c).value: c for c in range(1, 40) if ws.cell(1, c).value}
    rows = []
    seen: set[str] = set()
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, hdr["Name (trade goods)"]).value
        app = ws.cell(r, hdr.get("App template name", 2)).value
        canonical = str(name or app or "").strip()
        if not canonical or canonical in seen:
            continue
        seen.add(canonical)
        row = {"name": canonical}
        for h, c in hdr.items():
            row[h] = ws.cell(r, c).value
        row["name"] = canonical
        rows.append(row)
    return rows


def fx_args(row: dict) -> list[str]:
    parts = []
    for col, arg in PPB_COLS:
        v = pn(row.get(col))
        if v is None:
            continue
        if v == int(v):
            parts.append(f"{arg}: {int(v)}")
        else:
            parts.append(f"{arg}: {v}m")
    return parts


def generate_add(row: dict) -> str:
    name = row["name"]
    lordship = int(row.get("Lordship") or 1)
    kind = row.get("Kind") or "Building"
    kind_cs = "B" if kind == "Building" else "I"
    prod = int(pn(row.get("Production cost")) or 0)
    gold = int(pn(row.get("Gold cost")) or 0)
    notes = row.get("Notes")
    desc = str(notes).strip() if notes else f"{name}."
    desc_cs = esc(desc).replace("\n", "\\n")
    fx = fx_args(row)
    terrain = TERRAIN_BY_NAME.get(name)
    lines = []
    if fx:
        fx_s = ", ".join(fx)
        lines.append(f'            Add("{esc(name)}", {lordship}, {kind_cs}, {prod}, {gold},')
        lines.append(f'                "{desc_cs}",')
        if terrain:
            lines.append(f"                Fx({fx_s}),")
            lines.append(f'                terrainRequirement: "{terrain}");')
        else:
            lines.append(f"                Fx({fx_s}));")
    else:
        lines.append(f'            Add("{esc(name)}", {lordship}, {kind_cs}, {prod}, {gold},')
        lines.append(f'                "{desc_cs}");')
    return "\n".join(lines)


def split_seeder_shell(text: str) -> tuple[str, str]:
    """Return (prefix through const I, suffix from return list)."""
    m = re.search(
        r"const string B = BuildingKind\.Building;\n"
        r"            const string I = BuildingKind\.Improvement;\n\n",
        text,
    )
    if not m:
        raise RuntimeError("Could not find kind constants")
    prefix = text[: m.end()]
    suffix_m = re.search(r"\n            return list;\n        \}\n    \}\n\}\s*$", text[m.end() :])
    if not suffix_m:
        raise RuntimeError("Could not find return list")
    suffix = text[m.end() + suffix_m.start() :]
    return prefix, suffix


def main() -> None:
    excel_rows = load_excel_buildings()
    text = SEEDER.read_text(encoding="utf-8")
    prefix, suffix = split_seeder_shell(text)

    body = "\n\n".join(generate_add(r) for r in excel_rows) + "\n"
    SEEDER.write_text(prefix + body + suffix, encoding="utf-8")
    print(f"Updated {SEEDER}")
    print(f"  Excel buildings written: {len(excel_rows)}")


if __name__ == "__main__":
    main()

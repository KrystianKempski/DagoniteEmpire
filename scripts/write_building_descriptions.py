#!/usr/bin/env python3
"""Write rich building descriptions into Buildings sheet Notes, then sync seeder."""

from __future__ import annotations

from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs/barony/Katalog_towarow_handlowych.xlsx"

# name -> (body paragraphs joined by space/newline, produces line or None)
DESCRIPTIONS: dict[str, tuple[str, str | None]] = {
    "Amber gatherer": (
        "Coastal gatherers and beachcombers collect fossil resin washed up by storms and tides. "
        "The work is seasonal and exposed, but amber sells well to jewelers and temple craftsmen.",
        "Produces Amber",
    ),
    "Apiary": (
        "Beehives set among flowering meadows or forest clearings. "
        "Beekeepers harvest honey for food and trade, and beeswax for candles, seals, and crafts.",
        "Produces Honey & Wax",
    ),
    "Candlemaker": (
        "A workshop that renders beeswax and tallow into candles for homes, halls, and shrines. "
        "It does not create a trade good by itself, but turns Honey & Wax into everyday light and ritual goods.",
        None,
    ),
    "Armorer": (
        "A specialized workshop fitting mail, brigandine, and cuirasses. "
        "Needs iron or dagoferrite (and leather for straps and padding).",
        "Produces Medium Armor",
    ),
    "Plate Workshop": (
        "A master plate shop and armorers' guild hall for full and partial plate harness. "
        "Requires iron or dagoferrite and skilled plate-workers.",
        "Produces Heavy Armor",
    ),
    "Smithy": (
        "A small forge for tools, fittings, and simple weapons when soft metals or iron are available. "
        "Can be expanded into a full Forge for military-grade arms.",
        "Produces Simple Weapons",
    ),
    "Forge": (
        "A large military forge for swords, polearms, and war-grade blades. "
        "Requires iron or dagoferrite in the supply chain.",
        "Produces Military Weapons",
    ),
    "Brewery": (
        "A town brewery that malts grain into ale and beer for markets, inns, and garrisons. "
        "Reliable drink lifts spirits and keeps the common folk loyal when harvests are thin.",
        "Produces Beer",
    ),
    "Brickyard": (
        "Kilns and drying yards that fire clay into durable brick for walls, chimneys, and paved courts. "
        "Requires clay from a pit; finished brick opens masonry works that timber alone cannot match.",
        "Produces Bricks",
    ),
    "Cheese dairy": (
        "A dairy house where milk from cattle herds is turned into cheese for storage and sale. "
        "Needs cattle nearby; the dairy adds culture and comfort beyond raw meat and hides.",
        "Produces Cheese",
    ),
    "Clay pit": (
        "Open pits where laborers dig clay for bricks, tiles, and pottery. "
        "Hard outdoor work that scars the land, but supplies the raw material every ceramic craft depends on.",
        "Produces Clay",
    ),
    "Cotton farm": (
        "Warm-climate fields growing cotton for yarn and cloth. "
        "Needs fertile ground and heat; the crop feeds weavers and dyers once harvested.",
        "Produces Cotton",
    ),
    "Woad farm": (
        "Fields of woad (Isatis tinctoria) grown for blue dye. "
        "Leaves are harvested and fermented for the vat; unlocks the dyer's workshop.",
        "Produces Woad",
    ),
    "Madder farm": (
        "Beds of madder (Rubia tinctorum) grown for red dye. "
        "Roots are dug after a few seasons; unlocks the dyer's workshop.",
        "Produces Madder",
    ),
    "Weld farm": (
        "Plots of weld (Reseda luteola) grown for yellow dye. "
        "Flowering stalks are cut and dried for the mordant bath; unlocks the dyer's workshop.",
        "Produces Weld",
    ),
    "Farm (Dye plant)": (
        "Fields planted with the medieval dye triad—woad (blue), madder (red), and weld (yellow). "
        "Any of the three crops unlocks a dyer's workshop and feeds its vats.",
        "Produces Woad, Madder & Weld",
    ),
    "Dyer's workshop": (
        "Vats and drying racks for coloring cloth, banners, and uniforms. "
        "The smell and runoff bother neighbors, but dyed goods command far better prices than plain weave.",
        "Produces Dyes",
    ),
    "Flax farm": (
        "Fields of flax (and often hemp nearby) retted and dressed into fiber for linen, rope, and sails. "
        "A sturdy rural improvement that feeds the weaver's workshop and shipyards alike.",
        "Produces Flax & Hemp",
    ),
    "Glassworks": (
        "A hot furnace turning sand and cullet into window glass, vessels, and lenses. "
        "Costly to build and fuel, but glass marks a town as cultured and useful to scholars.",
        "Produces Glass",
    ),
    "Gunsmith workshop": (
        "A rare, expensive shop for muskets, arquebuses, and powder fittings. "
        "Needs sulfur and saltpeter in the supply chain; once running, it arms units with powder weapons.",
        "Produces Firearms",
    ),
    "Herbalist / Alchemist's Workshop": (
        "A herbalist's bench and simple alchemist's gear for salves, simples, and field reagents. "
        "Draws on dense woods or gardens for roots and herbs that heal, research, and trade.",
        "Produces Herbs & Roots",
    ),
    "Meat saltery (salt + fish/meat)": (
        "A salting house where fish and meat are cured for winter stores and long-distance carts. "
        "Needs salt plus a catch or slaughter; the result keeps far longer than fresh flesh.",
        "Produces Salted Fish & Meat",
    ),
    "Mine - Dagoferryt": (
        "A hard, dangerous mine for dagoferryt—the rare metal prized for the finest blades and armor. "
        "Extraction is costly, but the ore unlocks good-quality weapons beyond ordinary ironwork.",
        "Produces Dagoferryt",
    ),
    "Mine - Gold": (
        "Deep workings that yield gold for coin, tribute, and prestige. "
        "The ore also feeds a jeweler's bench: finished ornaments and regalia turn raw bullion into culture and court display.",
        "Produces Gold",
    ),
    "Mine - Iron": (
        "Iron pits and bloomery work that feed forges across the barony. "
        "Without iron there is no military smithing, blacksmith workshop expansion, or solid tools for builders.",
        "Produces Iron",
    ),
    "Mine - precious gems (luxury)": (
        "A luxury dig for precious and ornamental stones. "
        "Greed follows the haul, but gems sold to jewelers and temples bring enormous treasury and prestige.",
        "Produces Gemstones",
    ),
    "Mine - Salt": (
        "Salt workings or brine pans that preserve food and underwrite long-distance trade. "
        "Salt unlocks salting houses and keeps armies and towns fed through lean seasons.",
        "Produces Salt",
    ),
    "Mine - Silver": (
        "Silver veins for coinage, plate, and jewelry. "
        "Bullion can go to a mint or to a jeweler for ornaments that raise a court's standing.",
        "Produces Silver",
    ),
    "Mine - soft metals": (
        "Workings for copper, tin, lead, and similar soft metals. "
        "Essential for tools, alloys, fittings, and simple poor-quality weapons when iron is scarce.",
        "Produces Soft Metals (Cu, Sn, Pb)",
    ),
    "Mine - Sulfur": (
        "A foul sulfur mine—bad air and brittle ground, but vital for alchemy and gunpowder. "
        "Together with saltpeter it unlocks firearms and powder weapons for the army.",
        "Produces Sulfur",
    ),
    "Paper mill": (
        "Beaters and drying lofts that turn pulp into paper for ledgers, letters, and scriptoria. "
        "Supports administration, scholarship, and intelligence work far beyond parchment alone.",
        "Produces Paper",
    ),
    "Pastures (cattle)": (
        "Open grazing for cattle herds that supply meat, hides, and draft strength. "
        "Gold cost is high because the herd itself must be bought and stocked—pasture alone does not create cattle. "
        "Requires trade access to Cattle (breeding stock via treaty, import, or MG). Once built, this site becomes a local source.",
        "Produces Cattle",
    ),
    "Potter's workshop": (
        "Wheels and kilns for pots, tiles, and glazed wares from local clay. "
        "Everyday ceramics improve household life and local trade without needing imported luxuries.",
        "Produces Ceramics",
    ),
    "Quarry - common stone": (
        "A quarry for ordinary building stone—walls, foundations, and roads. "
        "Not as hard as granite, but enough for larger buildings and town defenses.",
        "Produces Building Stone",
    ),
    "Quarry - Granite": (
        "Hard granite cut for towers, keeps, and serious fortification. "
        "Slow and heavy work, but the stone shrugs off siege better than timber or soft rock.",
        "Produces Granite",
    ),
    "Quarry - Obsidian": (
        "A volcanic-glass quarry for sharp tools, ritual blades, and rare craft. "
        "Dangerous footing and scarce deposits, yet prized by smiths and mystics alike.",
        "Produces Obsidian",
    ),
    "Quarry - Tarnit": (
        "Extraction of tarnit, the heaviest and finest stone for top-tier fortresses. "
        "Enormous effort and cost, but nothing else matches it for the strongest walls.",
        "Produces Tarnit",
    ),
    "Saltpeter works": (
        "Beds and sheds where saltpeter is scraped and refined near towns. "
        "Unpleasant neighbors and poor loyalty, but essential—with sulfur—for firearms and powder.",
        "Produces Saltpeter",
    ),
    "Sawmill - Elven alder": (
        "A forest mill cutting rare elven alder—superb for bows and luxury carving. "
        "Scarce timber brings high prices, culture, and intrigue wherever it is traded.",
        "Produces Elven Alder",
    ),
    "Sawmill - Ironwood": (
        "Millwork on ironwood: dense, rare trunks that yield superior shafts and structures. "
        "Hard on saws and crews, but the timber supports good-quality weapons and elite crafts.",
        "Produces Ironwood",
    ),
    "Sawmill - Shipbuilding wood": (
        "A coastal or river mill dressing dense ship timber for hulls, masts, and yards. "
        "Without it, shipyards cannot lay proper keels or expand fleets.",
        "Produces Shipbuilding Timber",
    ),
    "Sheep pastures": (
        "Pastures stocked with sheep for wool, meat, and manure on milder slopes. "
        "Cheaper than cattle herds, yet the fleece feeds the entire cloth chain. "
        "Requires trade access to Sheep (breeding stock via treaty, import, or MG). Once built, this site becomes a local source.",
        "Produces Sheep & Wool",
    ),
    "Horse Stud (regular)": (
        "A stud farm breeding ordinary riding stock for patrols and light cavalry. "
        "Gold cost covers buying and stocking the herd—pasture alone does not create horses. "
        "Requires trade access to Horses (breeding stock via treaty, import, or MG). Once built, this site becomes a local source.",
        "Produces Horses",
    ),
    "Horse Stud (military)": (
        "A war-horse stud training chargers for medium and heavy cavalry. "
        "Demanding feed and handlers; mounts must be bred and drilled for battle. "
        "Requires trade access to War horses (breeding stock via treaty, import, or MG). Once built, this site becomes a local source.",
        "Produces War Horses",
    ),
    "Horse Stud (noble)": (
        "A prestigious stud for show and tourney breeds favored by the richest lords. "
        "Costly to maintain, but the mounts bring culture and prestige. "
        "Requires trade access to Noble horses (breeding stock via treaty, import, or MG). Once built, this site becomes a local source.",
        "Produces Noble Horses",
    ),
    "Tannery": (
        "A foul but necessary yard where hides become leather for boots, straps, and light armor. "
        "The stench hurts stability and law, yet armies and crafts cannot do without it.",
        "Produces Leather & Light Armor",
    ),
    "Vineyard": (
        "Terraced vines on warm hills, tended for wine, feasts, and export. "
        "Production cost is high because vines take years to mature before they give a real harvest.",
        "Produces Wine",
    ),
    "Weaver's workshop": (
        "Looms that turn flax, wool, or cotton into finished cloth for clothing and trade. "
        "The heart of the textile chain—feeds dyers and raises both economy and culture.",
        "Produces Cloth",
    ),
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Buildings"]
    hdr = {ws.cell(1, c).value: c for c in range(1, 40) if ws.cell(1, c).value}
    notes_col = hdr["Notes"]
    name_col = hdr["Name (trade goods)"]

    missing = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        if not name:
            continue
        name = str(name).strip()
        if name not in DESCRIPTIONS:
            missing.append(name)
            continue
        body, produces = DESCRIPTIONS[name]
        text = body.strip()
        if produces:
            text = f"{text}\n{produces}"
        ws.cell(r, notes_col).value = text
        print(f"OK {name}")

    wb.save(XLSX)
    if missing:
        print("Missing descriptions:", missing)
    print(f"Wrote notes for {len(DESCRIPTIONS)} buildings")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
Update docs/barony/Katalog_towarow_handlowych.xlsx (English sheets) and optionally
regenerate TradeGoodsCatalog.cs from sheet "All goods".

Run from repo: python3 DagoniteEmpire/scripts/sync_trade_goods_workbook.py [--write-cs]
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs/barony/Katalog_towarow_handlowych.xlsx"
SEEDER = ROOT / "DagoniteEmpire/Service/BuildingTemplateSeeder.cs"
CATALOG_CS = ROOT / "DA_Common/Barony/TradeGoodsCatalog.cs"

PPB_COLS = [
    "Food", "Economy", "Production", "Loyalty", "Stability", "Law", "Corruption",
    "Science", "Magic", "Culture", "Intelligence", "Defense", "Treasury",
]

# Canonical goods (Excel / app source of truth). Order matches design workbook.
GOODS: list[dict] = [
    {"key": "salt", "category": "Mines & stone", "name": "Salt",
     "description": "Salt from mines or brine works; preserves food and supports long-distance trade.",
     "bonus": "+5% Food", "unlocks": "Salted fish and meat", "building": "Mine - Salt", "requires": "salt deposit"},
    {"key": "fish-meat-salted", "category": "Food", "name": "Salted fish & meat",
     "description": "Fish and meat cured with salt; keeps well for storage and export.",
     "bonus": "+1 Defense, +1 Production", "unlocks": "", "building": "Meat saltery (salt + fish/meat)", "requires": "Salt + fish"},
    {"key": "cattle", "category": "Food", "name": "Cattle",
     "description": "Herds on pasture; meat, hides, wool, and draft power.",
     "bonus": "+1 Economy, +1 Production", "unlocks": "Meat, hides, cheese", "building": "Pastures (cattle)", "requires": "fertile land (3+)"},
    {"key": "cheese", "category": "Food", "name": "Cheese",
     "description": "Dairy from pastoral herds; staple food and local trade.",
     "bonus": "+1 Culture, +1 Economy", "unlocks": "", "building": "Cheese dairy", "requires": "cattle"},
    {"key": "honey-wax", "category": "Wood & forest", "name": "Honey & wax",
     "description": "Honey and beeswax from apiaries; sweetener, candles, and crafts.",
     "bonus": "+1 Loyalty, +1 Science", "unlocks": "", "building": "Apiary", "requires": "rich forest (3+)"},
    {"key": "wine", "category": "Food", "name": "Wine",
     "description": "Wine from vineyards; drink, feasts, and trade in warm hills.",
     "bonus": "+1 Intelligence, +1 Stability", "unlocks": "", "building": "Vineyard", "requires": "fertile hills, warm climate (4+)"},
    {"key": "olive-oil", "category": "Exotic imports", "name": "Olive oil",
     "description": "Cooking oil from olives; lamps and preserved foods.",
     "bonus": "+1 Culture, +1 Stability", "unlocks": "", "building": "Import", "requires": ""},
    {"key": "beer", "category": "Food", "name": "Beer",
     "description": "Ale and beer from grain; everyday drink in towns and garrisons.",
     "bonus": "+1 Stability, +1 Loyalty", "unlocks": "", "building": "Brewery", "requires": "grain"},
    {"key": "wool", "category": "Fibers & hides", "name": "Wool",
     "description": "Raw wool from sheep; spun into yarn and cloth.",
     "bonus": "+1 Production, +1 Economy", "unlocks": "Cloth", "building": "Sheep pastures", "requires": "fertile terrain (2+)"},
    {"key": "flax-hemp", "category": "Fibers & hides", "name": "Flax & hemp",
     "description": "Fiber for linen, rope, and sails.",
     "bonus": "+1 Production, +1 Stability", "unlocks": "Cloth", "building": "Flax farm", "requires": "fertile terrain (2+)"},
    {"key": "cloth", "category": "Fibers & hides", "name": "Cloth",
     "description": "Finished cloth for clothing, uniforms, and export.",
     "bonus": "+1 Economy, +1 Culture", "unlocks": "Dyed cloth", "building": "Weaver's workshop", "requires": "flax / wool / cotton"},
    {"key": "cotton", "category": "Fibers & hides", "name": "Cotton",
     "description": "Raw cotton; grows in warm, fertile fields.",
     "bonus": "+1 Production, +1 Loyalty", "unlocks": "Dyed cloth", "building": "Cotton farm", "requires": "fertile land (3+), warm climate"},
    {"key": "leather", "category": "Fibers & hides", "name": "Leather",
     "description": "Tanned hides for footwear, straps, and light armor.",
     "bonus": "+1 Production, +1 Defense", "unlocks": "Light armor", "building": "Tannery", "requires": "cattle"},
    {"key": "furs", "category": "Fibers & hides", "name": "Furs",
     "description": "Furs from hunts and cold country; luxury and winter wear.",
     "bonus": "+1 Culture, +1 Intelligence", "unlocks": "", "building": "Import", "requires": ""},
    {"key": "dyes", "category": "Craft goods", "name": "Dyes",
     "description": "Dyes and fixatives for cloth and banners.",
     "bonus": "+1 Culture, +1 Production", "unlocks": "Dyed cloth", "building": "Dyer's workshop", "requires": "dye source (woad etc.)"},
    {"key": "silk", "category": "Fibers & hides", "name": "Silk",
     "description": "Silk thread and fabric; expensive luxury import.",
     "bonus": "+2 Culture", "unlocks": "", "building": "Import", "requires": ""},
    {"key": "soft-metals", "category": "Mines & stone", "name": "Soft metals (Cu, Sn, Pb)",
     "description": "Copper, tin, and lead for tools, alloys, and simple weapons.",
     "bonus": "+1 Production", "unlocks": "Simple weapons (poor quality)", "building": "Mine - soft metals", "requires": "soft metals deposit"},
    {"key": "iron", "category": "Mines & stone", "name": "Iron",
     "description": "Iron ore and blooms; foundation of tools and military arms.",
     "bonus": "+1 Production, +1 Defense", "unlocks": "Military weapons, normal quality, Blacksmith Workshop", "building": "Mine - Iron", "requires": "iron deposit"},
    {"key": "silver", "category": "Mines & stone", "name": "Silver",
     "description": "Silver for coinage, jewelry, and treasury.",
     "bonus": "+1 Economy, +1 Culture", "unlocks": "Mint, jeweler", "building": "Mine - Silver", "requires": "silver deposit"},
    {"key": "gold", "category": "Mines & stone", "name": "Gold",
     "description": "Precious metal for coins, regalia, and prestige.",
     "bonus": "+1 Economy, +1 Culture", "unlocks": "Mint, jeweler", "building": "Mine - Gold", "requires": "gold deposit"},
    {"key": "gemstones", "category": "Mines & stone", "name": "Gemstones",
     "description": "Precious and ornamental stones.",
     "bonus": "+1 Culture, +1 Magic", "unlocks": "Jeweler", "building": "Mine - precious gems (luxury)", "requires": ""},
    {"key": "building-stone", "category": "Mines & stone", "name": "Building stone",
     "description": "Common stone for walls, houses, and roads.",
     "bonus": "+1 Production, +1 Defense", "unlocks": "Larger buildings, walls", "building": "Quarry - common stone", "requires": "stone deposit"},
    {"key": "granite", "category": "Mines & stone", "name": "Granite",
     "description": "Hard stone for towers and fortress walls.",
     "bonus": "+2 Defense", "unlocks": "Keeps, towers, defensive walls", "building": "Quarry - Granite", "requires": "granite deposit"},
    {"key": "clay", "category": "Mines & stone", "name": "Clay",
     "description": "Clay from pits; raw material for bricks and pottery.",
     "bonus": "+1 Production, +1 Economy", "unlocks": "Pottery workshop, brickyard", "building": "Clay pit", "requires": "clay deposit"},
    {"key": "bricks", "category": "Craft goods", "name": "Bricks",
     "description": "Fired brick for durable masonry.",
     "bonus": "+1 Production, +1 Defense", "unlocks": "Masonry buildings and walls", "building": "Brickyard", "requires": "clay"},
    {"key": "ceramics", "category": "Craft goods", "name": "Ceramics",
     "description": "Pots, tiles, and glazed wares.",
     "bonus": "+1 Culture, +1 Loyalty", "unlocks": "", "building": "Potter's workshop", "requires": "clay"},
    {"key": "shipbuilding-wood", "category": "Wood & forest", "name": "Shipbuilding timber",
     "description": "Timber for hulls, masts, and shipyards.",
     "bonus": "+1 Production, +1 Defense", "unlocks": "Shipyards", "building": "Sawmill - Shipbuilding wood", "requires": "shipbuilding wood"},
    {"key": "ironwood", "category": "Wood & forest", "name": "Ironwood",
     "description": "Very hard, rare timber for superior crafts and structures.",
     "bonus": "+1 Production, +1 Defense", "unlocks": "Weapons (good quality)", "building": "Sawmill - Ironwood", "requires": "ironwood forest"},
    {"key": "elven-alder", "category": "Wood & forest", "name": "Elven alder",
     "description": "Rare bow- and luxury-grade wood from deep forests.",
     "bonus": "+1 Culture, +1 Magic", "unlocks": "", "building": "Sawmill - Elven alder", "requires": "elven alder"},
    {"key": "herbs-roots", "category": "Wood & forest", "name": "Herbs & roots",
     "description": "Medicinal and alchemical plants from woods and gardens.",
     "bonus": "+1 Science, +1 Magic", "unlocks": "", "building": "Herbalist / Alchemist's Workshop", "requires": "dense forest"},
    {"key": "paper", "category": "Craft goods", "name": "Paper",
     "description": "Paper for records, letters, and scriptoria.",
     "bonus": "+1 Science, +1 Intelligence", "unlocks": "Scriptorium", "building": "Paper mill", "requires": ""},
    {"key": "glass", "category": "Craft goods", "name": "Glass",
     "description": "Glass for windows, vessels, and instruments.",
     "bonus": "+1 Culture, +1 Science", "unlocks": "", "building": "Glassworks", "requires": "sand/gravel pit"},
    {"key": "spices", "category": "Exotic imports", "name": "Spices",
     "description": "Pepper, ginger, cinnamon, and similar distant imports.",
     "bonus": "+2 Economy", "unlocks": "", "building": "Import", "requires": ""},
    {"key": "sugar", "category": "Exotic imports", "name": "Sugar",
     "description": "Cane sugar; costly sweetener and preserves.",
     "bonus": "+1 Culture, +1 Loyalty", "unlocks": "", "building": "Import", "requires": ""},
    {"key": "amber", "category": "Craft goods", "name": "Amber",
     "description": "Fossil resin from shores; jewelry and charms.",
     "bonus": "+1 Culture, +1 Magic", "unlocks": "Jeweler", "building": "Amber gatherer", "requires": "amber coast"},
    {"key": "ivory", "category": "Exotic imports", "name": "Ivory & walrus bone",
     "description": "Carving stock for luxury goods and ornament.",
     "bonus": "+1 Culture, +1 Production", "unlocks": "", "building": "Import", "requires": ""},
    {"key": "horses", "category": "Arms & horses", "name": "Horses",
     "description": "Riding stock and light cavalry mounts.",
     "bonus": "+1 Defense, +1 Production", "unlocks": "Light cavalry", "building": "Stables", "requires": "fertile land"},
    {"key": "war-horses", "category": "Arms & horses", "name": "War horses",
     "description": "Trained chargers for medium and heavy cavalry.",
     "bonus": "+2 Defense", "unlocks": "Medium and heavy cavalry", "building": "Stables", "requires": ""},
    {"key": "noble-horses", "category": "Arms & horses", "name": "Noble horses",
     "description": "Show and tourney breeds.",
     "bonus": "+1 Culture, +1 Loyalty", "unlocks": "", "building": "Stables", "requires": ""},
    {"key": "sulfur", "category": "Mines & stone", "name": "Sulfur",
     "description": "Sulfur from mines; gunpowder and alchemy.",
     "bonus": "+1 Production, +1% Food", "unlocks": "Firearms", "building": "Mine - Sulfur", "requires": "sulfur deposit"},
    {"key": "saltpeter", "category": "Craft goods", "name": "Saltpeter",
     "description": "Saltpeter for gunpowder; often produced near towns.",
     "bonus": "+1 Production, +1% Food", "unlocks": "Firearms", "building": "Saltpeter works", "requires": "town"},
    {"key": "access-arms-military", "category": "Arms & horses", "name": "Military weapons",
     "description": "Standing access to military-grade arms: swords, polearms, war bows.",
     "bonus": "+2 Defense", "unlocks": "Units: military weapons", "building": "Blacksmith Workshop", "requires": "iron"},
    {"key": "access-arms-firearms", "category": "Arms & horses", "name": "Firearms",
     "description": "Access to powder weapons: muskets and arquebuses.",
     "bonus": "+2 Defense", "unlocks": "Units: powder weapons", "building": "Gunsmith workshop", "requires": "sulfur, saltpeter"},
    {"key": "access-armor-light", "category": "Arms & horses", "name": "Light armor",
     "description": "Leather, caps, and padded protection.",
     "bonus": "+2 Defense", "unlocks": "Light armor", "building": "Tannery", "requires": "leather"},
    {"key": "access-armor-medium", "category": "Arms & horses", "name": "Medium armor",
     "description": "Mail, cuirasses, and brigandine.",
     "bonus": "+2 Defense", "unlocks": "Medium armor", "building": "Armorer (plate worker)", "requires": "leather, iron"},
    {"key": "access-armor-heavy", "category": "Arms & horses", "name": "Heavy armor",
     "description": "Plate and partial plate harness.",
     "bonus": "+2 Defense", "unlocks": "Heavy armor", "building": "Armorer (plate worker)", "requires": "leather, iron"},
    {"key": "obsidian", "category": "Mines & stone", "name": "Obsidian",
     "description": "Volcanic glass; sharp tools and ritual blades.",
     "bonus": "+1 Magic, +1 Production", "unlocks": "", "building": "Quarry - Obsidian", "requires": "obsidian deposit"},
    {"key": "tarnit", "category": "Mines & stone", "name": "Tarnit",
     "description": "Rare stone prized for the strongest fortifications.",
     "bonus": "+1 Magic, +1 Defense", "unlocks": "Top-tier fortifications", "building": "Quarry - Tarnit", "requires": "tarnit deposit"},
    {"key": "dagoferryt", "category": "Mines & stone", "name": "Dagoferryt",
     "description": "Rare metal for the finest weapons and armor.",
     "bonus": "+2 Defense", "unlocks": "Good-quality weapons", "building": "Mine - Dagoferryt", "requires": "dagoferryt deposit"},
    {"key": "elf-forest-goods", "category": "Exotic imports", "name": "Elven forest crafts",
     "description": "Finished luxury goods from distant elven woods.",
     "bonus": "+1 Culture, +1 Magic", "unlocks": "", "building": "Import", "requires": ""},
]

CATEGORY_TO_SECTION = {
    "Food": "Food",
    "Fibers & hides": "Fibers",
    "Mines & stone": "Mines",
    "Wood & forest": "Wood",
    "Craft goods": "Crafts",
    "Exotic imports": "Exotic",
    "Arms & horses": "Military",
}

# Trade-goods production label -> BuildingTemplateSeeder name (if any)
BUILDING_APP_ALIASES: dict[str, str] = {
    "Herbalist / Alchemist's Workshop": "Herbalist / Alchemist's Workshop",
    "Import": "",
}

PL_CATEGORY_EN = {
    "kopaliny": "Mines & stone",
    "żywność": "Food",
    "drewno i owoce lasu": "Wood & forest",
    "drewno I owoce lasu": "Wood & forest",
    "egzotyczne": "Exotic imports",
    "włókna i skóry": "Fibers & hides",
    "włókna I skóry": "Fibers & hides",
    "wyroby rzemieślnicze": "Craft goods",
    "broń i konie": "Arms & horses",
    "broń I konie": "Arms & horses",
}

# Polish production labels -> English (trade goods column)
BUILDING_PL_TO_EN: dict[str, str] = {
    "solernia mięsa (wymaga soli + mięso)": "Meat saltery (salt + fish/meat)",
    "pastwiska (krowy)": "Pastures (cattle)",
    "serowarnia": "Cheese dairy",
    "pasieka": "Apiary",
    "winnica": "Vineyard",
    "import": "Import",
    "browar": "Brewery",
    "pastwiska (owce)": "Sheep pastures",
    "farm (len)": "Flax farm",
    "warsztat tkacki": "Weaver's workshop",
    "farm (bawełna)": "Cotton farm",
    "garbarnia": "Tannery",
    "warsztat farbierski": "Dyer's workshop",
    "cegielnia": "Brickyard",
    "warsztat garncarski": "Potter's workshop",
    "warsztat papierniczy": "Paper mill",
    "huta szkła": "Glassworks",
    "zbieracz bursztynu": "Amber gatherer",
    "stadnina": "Stables",
    "saletrarnia": "Saltpeter works",
    "warsztat rusznikarza": "Gunsmith workshop",
    "płatnerz": "Armorer (plate worker)",
}

PL_PPB_WORDS = [
    (r"wyżywienie|żywności", "Food"),
    (r"ekonomia", "Economy"),
    (r"produkcja", "Production"),
    (r"lojalności|lojalność|lojalności", "Loyalty"),
    (r"stabilności|stlność|stabilność", "Stability"),
    (r"nauki|nauka", "Science"),
    (r"wywiad", "Intelligence"),
    (r"kultura", "Culture"),
    (r"magia", "Magic"),
    (r"obrony|obrona", "Defense"),
]

UNLOCKS_PL_TO_EN: dict[str, str] = {
    "solone ryby, solone mięso": "Salted fish & meat",
    "mięso, skóry, sery": "Meat, hides, cheese",
    "tkaniny": "Cloth",
    "tkaniny barwione": "Dyed cloth",
    "lekkie zbroje": "Light armor",
    "broń prosta, poor quality weapon": "Simple weapons (poor quality)",
    "broń wojskowa, normal quality weapon, blacksmith workshop": "Military weapons, normal quality, Blacksmith Workshop",
    "mennica, jubiler": "Mint, jeweler",
    "jubiler": "Jeweler",
    "większe budowle, mury": "Larger buildings, walls",
    "twierdze, wieże, mury obronne": "Keeps, towers, defensive walls",
    "warsztat ceramiczny, ceglarnia": "Pottery workshop, brickyard",
    "kamienne/ceglane budowle, mury": "Masonry buildings and walls",
    "stocznie": "Shipyards",
    "broń quality: good,": "Weapons (good quality)",
    "broń quality: good": "Weapons (good quality)",
    "skryptorium,": "Scriptorium",
    "skryptorium": "Scriptorium",
    "lekka kawaleria": "Light cavalry",
    "średnia i ciężka kawaleria": "Medium and heavy cavalry",
    "broń palna": "Firearms",
    "jednostki: broń wojskowa (military)": "Units: military weapons",
    "jednostki: broń prochowa (powder)": "Units: powder weapons",
    "zbroje lekkie": "Light armor",
    "zbroje średnie": "Medium armor",
    "zbroje ciężkie": "Heavy armor",
    "najlepsze fortyfikacje": "Top-tier fortifications",
    "dobra jakość broni": "Good-quality weapons",
}

PL_GOOD_TO_KEY: dict[str, str] = {
    "sól": "salt",
    "solone ryby/mięso": "fish-meat-salted",
    "bydło": "cattle",
    "sery": "cheese",
    "miód i wosk": "honey-wax",
    "wino": "wine",
    "oliwa": "olive-oil",
    "piwo": "beer",
    "wełna": "wool",
    "len i konopia": "flax-hemp",
    "tkaniny": "cloth",
    "bawełna": "cotton",
    "skóry": "leather",
    "futra": "furs",
    "barwniki": "dyes",
    "jedwab": "silk",
    "metale miękkie (miedź, cyna, ołów)": "soft-metals",
    "żelazo": "iron",
    "srebro": "silver",
    "złoto": "gold",
    "klejnoty": "gemstones",
    "kamień budowlany": "building-stone",
    "granit": "granite",
    "glina": "clay",
    "cegły": "bricks",
    "ceramika": "ceramics",
    "drewno okrętowe": "shipbuilding-wood",
    "żelazodrzewo": "ironwood",
    "elfia olcha": "elven-alder",
    "zioła i korzenie": "herbs-roots",
    "papier": "paper",
    "szkło": "glass",
    "przyprawy": "spices",
    "cukier": "sugar",
    "bursztyn": "amber",
    "kość słoniowa / morświna": "ivory",
    "konie": "horses",
    "konie bojowe": "war-horses",
    "konie arystokratyczne": "noble-horses",
    "siarka": "sulfur",
    "saletra": "saltpeter",
    "broń wojskowa (dostęp)": "access-arms-military",
    "broń palna (dostęp)": "access-arms-firearms",
    "lekkie zbroje (dostęp)": "access-armor-light",
    "średnie zbroje (dostęp)": "access-armor-medium",
    "ciężkie zbroje (dostęp)": "access-armor-heavy",
    "obsydian": "obsidian",
    "tarnit": "tarnit",
    "dagoferryt": "dagoferryt",
    "wyroby leśnych elfów": "elf-forest-goods",
}


def _norm_key(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())


def translate_bonus_pl(pl: str | None) -> str:
    if not pl:
        return ""
    s = str(pl).strip()
    s = re.sub(r"\+\s*kultura", "+1 Culture", s, flags=re.I)
    s = re.sub(r"\+\s*ekonomia", "+1 Economy", s, flags=re.I)
    s = re.sub(r"\+\s*1\s+", "+1 ", s)
    for pat, en in PL_PPB_WORDS:
        s = re.sub(pat, en, s, flags=re.I)
    s = re.sub(r"(\+\d+(?:\.\d+)?\s+[A-Za-z]+)\s+(\+\d)", r"\1, \2", s)
    s = re.sub(r",\s*$", "", s)
    return s


def translate_building_pl(pl: str | None) -> str:
    if not pl:
        return "Import"
    raw = normalize_building_label(str(pl))
    low = _norm_key(raw)
    if low in BUILDING_PL_TO_EN:
        return BUILDING_PL_TO_EN[low]
    if low == "import":
        return "Import"
    return raw


def translate_requires_pl(pl: str | None) -> str:
    if pl is None:
        return ""
    s = str(pl).strip().replace("\xa0", " ")
    if not s:
        return ""
    repl = {
        "sól": "salt deposit",
        "sól + ryby": "salt deposit + fish",
        "ryby": "fish",
        "żelazo": "iron",
        "srebro": "silver deposit",
        "złoto": "gold deposit",
        "kamień": "stone deposit",
        "granit": "granite deposit",
        "siarka": "sulfur",
        "saletra": "saltpeter",
        "siarka, saletra": "sulfur, saltpeter",
        "obsidian": "obsidian deposit",
        "tarnit": "tarnit deposit",
        "dagoferryt": "dagoferryt deposit",
        "bursztyn": "amber coast",
        "żyzne ziemie": "fertile land",
        "żyzne ziemie (3+)": "fertile land (3+)",
        "żyzne tereny (2+)": "fertile terrain (2+)",
        "żyzne wzgórza, ciepły klimat (4+)": "fertile hills, warm climate (4+)",
        "żyzne ziemie (3+), ciepły klimat": "fertile land (3+), warm climate",
        "żyzne lasy (3+)": "rich forest (3+)",
        "gęsty las": "dense forest",
        "żwirownia": "sand/gravel pit",
        "bydło": "cattle",
        "zboże": "grain",
        "skóra": "leather",
        "skóra, żelazo": "leather, iron",
        "miasto": "town",
        "clay": "clay deposit",
        "len/wełna/bawełna": "flax / wool / cotton",
        "żródło bawnika": "dye source (woad etc.)",
        "źródło barwnika": "dye source (woad etc.)",
        "źródło bawnika": "dye source (woad etc.)",
    }
    norm_map = {_norm_key(k): v for k, v in repl.items()}
    key = _norm_key(s)
    if key in norm_map:
        return norm_map[key]
    if re.search(r"\s+\+\s+", s):
        parts = re.split(r"\s+\+\s+", s)
        return " + ".join(translate_requires_pl(p) for p in parts)
    if "," in s:
        parts = [p.strip() for p in s.split(",")]
        return ", ".join(translate_requires_pl(p) for p in parts)
    return s


def import_goods_from_polish_sheet(wb: openpyxl.Workbook) -> tuple[list[dict], list[str]]:
    """Merge Polish 'Wszystkie towary' into canonical English goods rows."""
    ws = wb["Wszystkie towary"]
    canon = {g["key"]: dict(g) for g in GOODS}
    warnings: list[str] = []
    merged: list[dict] = []
    for r in range(3, ws.max_row + 1):
        cat_pl = ws.cell(r, 1).value
        name_pl = ws.cell(r, 2).value
        if not name_pl:
            continue
        key = PL_GOOD_TO_KEY.get(_norm_key(str(name_pl)))
        if not key:
            warnings.append(f"Row {r}: unknown good name {name_pl!r}")
            continue
        if key not in canon:
            warnings.append(f"Row {r}: key {key} not in catalog")
            continue
        g = dict(canon[key])
        if cat_pl:
            cat_en = PL_CATEGORY_EN.get(_norm_key(str(cat_pl)))
            if cat_en:
                g["category"] = cat_en
            else:
                warnings.append(f"Row {r}: unknown category {cat_pl!r}")
        bonus_pl = ws.cell(r, 4).value
        unlocks_pl = ws.cell(r, 5).value
        building_pl = ws.cell(r, 6).value
        req_pl = ws.cell(r, 7).value
        g["bonus"] = translate_bonus_pl(bonus_pl) if bonus_pl else g["bonus"]
        if unlocks_pl:
            raw = str(unlocks_pl).strip()
            # Normalize NBSP etc. before lookup
            key = _norm_key(raw.replace("\xa0", " "))
            g["unlocks"] = UNLOCKS_PL_TO_EN.get(key, raw)
        g["building"] = translate_building_pl(building_pl)
        g["requires"] = translate_requires_pl(req_pl)
        merged.append(g)
    if len(merged) != len(GOODS):
        warnings.append(f"Imported {len(merged)} goods, expected {len(GOODS)}")
    return merged, warnings


def parse_seeder_buildings() -> dict[str, dict]:
    text = SEEDER.read_text(encoding="utf-8")
    names = re.findall(r'Add\("((?:[^"\\]|\\.)*)"', text)
    # Parse Fx(...) blocks per Add — simplified: extract Fx lines after each Add
    buildings: dict[str, dict] = {}
    chunks = re.split(r'\n\s*Add\("', text)
    for chunk in chunks[1:]:
        m = re.match(r'((?:[^"\\]|\\.)*)"\s*,\s*(\d+)\s*,\s*([BI])\s*,', chunk)
        if not m:
            continue
        name = m.group(1).replace('\\"', '"')
        lordship = int(m.group(2))
        kind = "Building" if m.group(3) == "B" else "Improvement"
        prod_m = re.search(
            r'(?:,\s*([\d.]+)\s*,\s*([\d.]+)(?:\s*,|\s*\))|additive:\s*Fx)',
            chunk[:800],
        )
        production = gold = 0.0
        if prod_m and prod_m.group(1):
            production = float(prod_m.group(1))
            gold = float(prod_m.group(2))
        fx_m = re.search(r'Fx\(([^)]*)\)', chunk)
        ppb = {k: 0.0 for k in PPB_COLS}
        if fx_m and fx_m.group(1).strip():
            for part in re.findall(r'(\w+)\s*:\s*(-?[\d.]+)m?', fx_m.group(1)):
                key = part[0][0].upper() + part[0][1:] if part[0] == "food" else (
                    "Economy" if part[0] == "economy" else
                    "Production" if part[0] == "production" else
                    "Loyalty" if part[0] == "loyalty" else
                    "Stability" if part[0] == "stability" else
                    "Law" if part[0] == "law" else
                    "Corruption" if part[0] == "corruption" else
                    "Science" if part[0] == "science" else
                    "Magic" if part[0] == "magic" else
                    "Culture" if part[0] == "culture" else
                    "Intelligence" if part[0] == "intelligence" else
                    "Defense" if part[0] == "defense" else
                    "Treasury" if part[0] == "treasury" else part[0]
                )
                if key in ppb:
                    ppb[key] = float(part[1])
        buildings[name] = {
            "lordship": lordship,
            "kind": kind,
            "production_cost": production,
            "gold_cost": gold,
            "ppb": ppb,
        }
    return buildings


def normalize_building_label(label: str) -> str:
    return " ".join(label.split()).strip()


def resolve_app_building(label: str, app: dict[str, dict]) -> tuple[str | None, dict | None]:
    if not label or label.lower() == "import":
        return None, None
    label = normalize_building_label(label)
    if label in BUILDING_APP_ALIASES:
        alias = BUILDING_APP_ALIASES[label]
        if not alias:
            return None, None
        label = alias
    if label in app:
        return label, app[label]
    low = label.lower()
    for name in app:
        if name.lower() == low:
            return name, app[name]
    return None, None


def collect_building_rows(app: dict[str, dict], goods: list[dict] | None = None) -> list[dict]:
    source = goods if goods is not None else GOODS
    seen: dict[str, dict] = {}
    for g in source:
        b = normalize_building_label(g["building"])
        if not b or b.lower() == "import":
            continue
        if b in seen:
            continue
        app_name, data = resolve_app_building(b, app)
        row = {
            "name": b,
            "app_template": app_name or "",
            "in_app": "Y" if app_name else "N",
            "lordship": "",
            "kind": "",
            "production_cost": "",
            "gold_cost": "",
            **{c: "" for c in PPB_COLS},
            "notes": "",
        }
        if data:
            row["lordship"] = data["lordship"]
            row["kind"] = data["kind"]
            row["production_cost"] = data["production_cost"]
            row["gold_cost"] = data["gold_cost"]
            for k, v in data["ppb"].items():
                if v != 0:
                    row[k] = v
        else:
            row["notes"] = "Not in app BuildingTemplate catalog yet — fill PPB and costs here."
        seen[b] = row
    return sorted(seen.values(), key=lambda r: r["name"].lower())


def parse_bonus(bonus: str) -> tuple[dict[str, float], dict[str, float]]:
    add: dict[str, float] = {}
    pct: dict[str, float] = {}
    normalized = re.sub(r"\s+\+\s+", ", +", bonus.strip())
    for part in re.split(r",\s*", normalized):
        if not part:
            continue
        m = re.match(r"\+(\d+(?:\.\d+)?)\s*%\s*(.+)", part, re.I)
        if m:
            pct[m.group(2).strip()] = float(m.group(1))
            continue
        m = re.match(r"\+(\d+(?:\.\d+)?)\s+(.+)", part, re.I)
        if m:
            add[m.group(2).strip()] = float(m.group(1))
    return add, pct


def compare_polish_sheet(wb: openpyxl.Workbook) -> list[str]:
    issues: list[str] = []
    if "Wszystkie towary" not in wb.sheetnames:
        return issues
    ws = wb["Wszystkie towary"]
    pl_rows = []
    for r in range(3, ws.max_row + 1):
        cat, name = ws.cell(r, 1).value, ws.cell(r, 2).value
        if not name:
            continue
        pl_rows.append((str(name).strip(), ws.cell(r, 4).value, ws.cell(r, 6).value))
    if len(pl_rows) != len(GOODS):
        issues.append(f"Polish sheet row count {len(pl_rows)} vs canonical {len(GOODS)}")
    return issues


def write_workbook(
    path: Path,
    goods: list[dict] | None = None,
    *,
    preserve_goods: bool = False,
    keep_polish_sheet: openpyxl.Workbook | None = None,
) -> None:
    active_goods = goods if goods is not None else GOODS
    if preserve_goods and path.is_file():
        wb = openpyxl.load_workbook(path)
        if "All goods" in wb.sheetnames:
            refresh_buildings_sheet(wb)
            path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(path)
            return

    if keep_polish_sheet is not None:
        wb = keep_polish_sheet
        if "All goods" in wb.sheetnames:
            del wb["All goods"]
        if "Buildings" in wb.sheetnames:
            del wb["Buildings"]
        ws = wb.create_sheet("All goods", 0)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All goods"

    headers = [
        "Key", "Category", "Good", "Description", "PPB bonus", "Unlocks",
        "Production building", "Requires",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for g in active_goods:
        ws.append([
            g["key"], g["category"], g["name"], g["description"], g["bonus"],
            g["unlocks"], g["building"], g["requires"],
        ])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 36
    ws.column_dimensions["H"].width = 32

    app = parse_seeder_buildings()
    ws_b = wb.create_sheet("Buildings")
    b_headers = [
        "Name (trade goods)", "App template name", "In app catalog", "Lordship", "Kind",
        "Production cost", "Gold cost",
    ] + PPB_COLS + ["Notes"]
    ws_b.append(b_headers)
    for cell in ws_b[1]:
        cell.font = Font(bold=True)
    for row in collect_building_rows(app, active_goods):
        ws_b.append([
            row["name"], row["app_template"], row["in_app"], row["lordship"], row["kind"],
            row["production_cost"], row["gold_cost"],
        ] + [row[c] for c in PPB_COLS] + [row["notes"]])
    for col, w in zip("ABCDEFGHIJKLMNOPQRSTUVWXYZ", [32, 28, 12, 10, 14, 14, 12] + [9] * 13 + [48]):
        ws_b.column_dimensions[col].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def refresh_buildings_sheet(wb: openpyxl.Workbook) -> None:
    """Rebuild Buildings sheet from current 'All goods' rows + app seeder."""
    ws_g = wb["All goods"]
    goods_buildings: list[str] = []
    for r in range(2, ws_g.max_row + 1):
        b = ws_g.cell(r, 7).value
        if b:
            goods_buildings.append(normalize_building_label(str(b)))
    app = parse_seeder_buildings()
    seen: dict[str, dict] = {}
    for b in goods_buildings:
        if not b or b.lower() == "import":
            continue
        if b in seen:
            continue
        app_name, data = resolve_app_building(b, app)
        row = {
            "name": b,
            "app_template": app_name or "",
            "in_app": "Y" if app_name else "N",
            "lordship": "",
            "kind": "",
            "production_cost": "",
            "gold_cost": "",
            **{c: "" for c in PPB_COLS},
            "notes": "" if app_name else "Not in app BuildingTemplate catalog yet — fill PPB and costs here.",
        }
        if data:
            row["lordship"] = data["lordship"]
            row["kind"] = data["kind"]
            row["production_cost"] = data["production_cost"]
            row["gold_cost"] = data["gold_cost"]
            for k, v in data["ppb"].items():
                if v != 0:
                    row[k] = v
        seen[b] = row
    rows = sorted(seen.values(), key=lambda r: r["name"].lower())

    if "Buildings" in wb.sheetnames:
        del wb["Buildings"]
    ws_b = wb.create_sheet("Buildings")
    b_headers = [
        "Name (trade goods)", "App template name", "In app catalog", "Lordship", "Kind",
        "Production cost", "Gold cost",
    ] + PPB_COLS + ["Notes"]
    ws_b.append(b_headers)
    for cell in ws_b[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws_b.append([
            row["name"], row["app_template"], row["in_app"], row["lordship"], row["kind"],
            row["production_cost"], row["gold_cost"],
        ] + [row[c] for c in PPB_COLS] + [row["notes"]])


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--write-cs", action="store_true", help="Regenerate TradeGoodsCatalog.cs")
    ap.add_argument("--xlsx-only", action="store_true", help="Only update xlsx, skip TradeGoodsCatalog.cs")
    ap.add_argument("--refresh-buildings", action="store_true",
                    help="Rebuild Buildings sheet from All goods + app seeder (keeps goods sheet)")
    ap.add_argument("--from-polish", action="store_true",
                    help="Import goods from 'Wszystkie towary' (default if that sheet exists without All goods)")
    args = ap.parse_args()

    wb_full = openpyxl.load_workbook(XLSX) if XLSX.exists() else None
    use_polish = args.from_polish or (
        wb_full is not None
        and "Wszystkie towary" in wb_full.sheetnames
        and "All goods" not in wb_full.sheetnames
    )

    merged_goods: list[dict] | None = None
    import_warnings: list[str] = []
    if use_polish and wb_full is not None:
        merged_goods, import_warnings = import_goods_from_polish_sheet(wb_full)
        write_workbook(XLSX, merged_goods, keep_polish_sheet=wb_full)
        goods_count = len(merged_goods)
    elif args.refresh_buildings:
        if wb_full is None:
            print("Workbook missing", file=sys.stderr)
            return 1
        refresh_buildings_sheet(wb_full)
        wb_full.save(XLSX)
        print(f"Refreshed Buildings in {XLSX}")
        if not args.xlsx_only:
            import subprocess
            subprocess.check_call([sys.executable, str(ROOT / "scripts/gen_trade_goods_catalog.py")])
        return 0
    else:
        write_workbook(XLSX, preserve_goods=False)
        goods_count = len(GOODS)

    print(f"Updated {XLSX}")
    print(f"  Sheets: Wszystkie towary (kept), All goods ({goods_count} rows), Buildings")

    app = parse_seeder_buildings()
    rows = collect_building_rows(app, merged_goods)
    matched = sum(1 for r in rows if r["in_app"] == "Y")
    print(f"  Buildings: {len(rows)} unique, {matched} matched to app templates")

    for w in import_warnings:
        print(f"  ! {w}")

    if args.write_cs or not args.xlsx_only:
        import subprocess
        subprocess.check_call([sys.executable, str(ROOT / "scripts/gen_trade_goods_catalog.py")])
        print(f"Regenerated {CATALOG_CS}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

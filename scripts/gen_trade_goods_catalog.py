#!/usr/bin/env python3
"""Regenerate TradeGoodsCatalog.cs from docs/barony/Katalog_towarow_handlowych.xlsx (sheet All goods)."""

import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "DA_Common/Barony/TradeGoodsCatalog.cs"
XLSX = ROOT / "docs/barony/Katalog_towarow_handlowych.xlsx"

CATEGORY_TO_SECTION = {
    "Food": "Food",
    "Fibers & hides": "Fibers",
    "Mines & stone": "Mines",
    "Wood & forest": "Wood",
    "Craft goods": "Crafts",
    "Exotic imports": "Exotic",
    "Arms & horses": "Military",
}


def parse_bonus(bonus: str) -> tuple[dict[str, float], dict[str, float]]:
    add: dict[str, float] = {}
    pct: dict[str, float] = {}
    normalized = re.sub(r"\s+\+\s+", ", +", (bonus or "").strip())
    normalized = re.sub(r"(\+\d+(?:\.\d+)?\s+[A-Za-z]+)\s+(\+\d)", r"\1, \2", normalized)
    for part in re.split(r",\s*", normalized):
        if not part:
            continue
        m = re.match(r"\+(\d+(?:\.\d+)?)\s*%\s*(.+)", part, re.I)
        if m:
            pct[m.group(2).strip()] = float(m.group(1))
            continue
        m = re.match(r"\+(\d+(?:\.\d+)?)\s+(.+)", part, re.I)
        if m:
            add[m.group(2).strip()] = float(m.group(1))
    return add, pct


def load_goods_from_xlsx() -> list[tuple]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["All goods"]
    rows = []
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 1).value
        if not key:
            continue
        cat = str(ws.cell(r, 2).value or "").strip()
        sec = CATEGORY_TO_SECTION.get(cat)
        if not sec:
            raise ValueError(f"Row {r}: unknown category {cat!r}")
        name = str(ws.cell(r, 3).value or "")
        desc = str(ws.cell(r, 4).value or "")
        bonus = str(ws.cell(r, 5).value or "")
        unlocks = str(ws.cell(r, 6).value or "") if ws.cell(r, 6).value else ""
        building = str(ws.cell(r, 7).value or "")
        req_val = ws.cell(r, 8).value
        req = None if req_val is None or str(req_val).strip() == "" else str(req_val).strip()
        add, pct = parse_bonus(bonus)
        rows.append((str(key).strip(), sec, name, desc, bonus, unlocks, building, req, add, pct))
    return rows

HEADER = '''namespace DA_Common.Barony
{
    public static class TradeGoodSection
    {
        public const string Food = "Zywnosc";
        public const string Fibers = "Wloki";
        public const string Mines = "Kopaliny";
        public const string Wood = "Drewno";
        public const string Crafts = "Rzemioslo";
        public const string Exotic = "Egzotyczne";
        public const string Military = "Bron";

        public static readonly IReadOnlyList<(string Key, string LabelEn)> All = new (string, string)[]
        {
            (TradeGoodSection.Food, "Food"),
            (TradeGoodSection.Fibers, "Fibers & hides"),
            (TradeGoodSection.Mines, "Mines & stone"),
            (TradeGoodSection.Wood, "Wood & forest"),
            (TradeGoodSection.Crafts, "Craft goods"),
            (TradeGoodSection.Exotic, "Exotic imports"),
            (TradeGoodSection.Military, "Arms & horses"),
        };
    }

    public sealed class TradeGoodEntry
    {
        public required string Key { get; init; }
        public required string SectionKey { get; init; }
        public required string Name { get; init; }
        public required string Description { get; init; }
        public required string BonusDisplay { get; init; }
        public required string Unlocks { get; init; }
        public required string ProductionBuilding { get; init; }
        public string? Requirements { get; init; }
        public PpbVector BonusAdditive { get; init; } = new();
        public PpbVector BonusPercent { get; init; } = new();
    }

    public static class TradeGoodsCatalog
    {
        private static readonly List<TradeGoodEntry> _all = new();
        public static IReadOnlyList<TradeGoodEntry> All => _all;

        private static TradeGoodEntry G(
            string key, string section, string name, string description,
            string bonusDisplay, string unlocks, string productionBuilding, string? requirements)
        {
            var g = new TradeGoodEntry
            {
                Key = key, SectionKey = section, Name = name, Description = description,
                BonusDisplay = bonusDisplay, Unlocks = unlocks, ProductionBuilding = productionBuilding,
                Requirements = requirements,
            };
            _all.Add(g);
            return g;
        }

        static TradeGoodsCatalog()
        {
'''

FOOTER = '''
        }

        public static TradeGoodEntry? Find(string? key) =>
            string.IsNullOrWhiteSpace(key) ? null : _all.FirstOrDefault(x => string.Equals(x.Key, key, StringComparison.OrdinalIgnoreCase));

        public static IEnumerable<TradeGoodEntry> BySection(string sectionKey) =>
            _all.Where(x => x.SectionKey == sectionKey);
    }

    public static class TradeGoodsBonusAggregator
    {
        public static void Sum(IEnumerable<TradeGoodEntry> goods, out PpbVector additive, out PpbVector percent)
        {
            additive = new PpbVector();
            percent = new PpbVector();
            foreach (var g in goods)
            {
                additive.AddInPlace(g.BonusAdditive);
                percent.AddInPlace(g.BonusPercent);
            }
        }
    }
}
'''

# section, name, description, bonus_display, unlocks, building, requirements, additive dict, percent dict
# Ppb: Food, Economy, Culture, Defense, Production, Loyalty, Stability, Science, Magic, Intelligence

def esc(s):
    if s is None:
        return "null"
    return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'

# Fallback if xlsx missing (run sync_trade_goods_workbook.py to create it).
_GOODS_FALLBACK = [
    ('salt', 'Mines', 'Salt', 'Salt from mines or brine works; preserves food and supports long-distance trade.',
     '+5% Food', 'Salted fish and meat', 'Mine - Salt', 'salt deposit', {}, {"Food": 5.0}),
    ('fish-meat-salted', 'Food', 'Salted fish & meat', 'Fish and meat cured with salt; keeps well for storage and export.',
     '+1 Defense, +1 Production', '', 'Meat saltery (salt + fish/meat)', 'Salt + fish', {"Defense": 1.0, "Production": 1.0}, {}),
    ('cattle', 'Food', 'Cattle', 'Herds on pasture; meat, hides, wool, and draft power.',
     '+1 Economy, +1 Production', 'Meat, hides, cheese', 'Pastures (cattle)', 'fertile land (3+)', {"Economy": 1.0, "Production": 1.0}, {}),
    ('cheese', 'Food', 'Cheese', 'Dairy from pastoral herds; staple food and local trade.',
     '+1 Culture, +1 Economy', '', 'Cheese dairy', 'cattle', {"Culture": 1.0, "Economy": 1.0}, {}),
    ('honey-wax', 'Wood', 'Honey & wax', 'Honey and beeswax from apiaries; sweetener, candles, and crafts.',
     '+1 Loyalty, +1 Science', '', 'Apiary', 'rich forest (3+)', {"Loyalty": 1.0, "Science": 1.0}, {}),
    ('wine', 'Food', 'Wine', 'Wine from vineyards; drink, feasts, and trade in warm hills.',
     '+1 Intelligence, +1 Stability', '', 'Vineyard', 'fertile hills, warm climate (4+)', {"Intelligence": 1.0, "Stability": 1.0}, {}),
    ('olive-oil', 'Exotic', 'Olive oil', 'Cooking oil from olives; lamps and preserved foods.',
     '+1 Culture, +1 Stability', '', 'Import', None, {"Culture": 1.0, "Stability": 1.0}, {}),
    ('beer', 'Food', 'Beer', 'Ale and beer from grain; everyday drink in towns and garrisons.',
     '+1 Stability, +1 Loyalty', '', 'Brewery', 'grain', {"Stability": 1.0, "Loyalty": 1.0}, {}),
    ('wool', 'Fibers', 'Wool', 'Raw wool from sheep; spun into yarn and cloth.',
     '+1 Production, +1 Economy', 'Cloth', 'Sheep pastures', 'fertile terrain (2+)', {"Production": 1.0, "Economy": 1.0}, {}),
    ('flax-hemp', 'Fibers', 'Flax & hemp', 'Fiber for linen, rope, and sails.',
     '+1 Production, +1 Stability', 'Cloth', 'Flax farm', 'fertile terrain (2+)', {"Production": 1.0, "Stability": 1.0}, {}),
    ('cloth', 'Fibers', 'Cloth', 'Finished cloth for clothing, uniforms, and export.',
     '+1 Economy, +1 Culture', 'Dyed cloth', "Weaver's workshop", 'flax / wool / cotton', {"Economy": 1.0, "Culture": 1.0}, {}),
    ('cotton', 'Fibers', 'Cotton', 'Raw cotton; grows in warm, fertile fields.',
     '+1 Production, +1 Loyalty', 'Dyed cloth', 'Cotton farm', 'fertile land (3+), warm climate', {"Production": 1.0, "Loyalty": 1.0}, {}),
    ('leather', 'Fibers', 'Leather', 'Tanned hides for footwear, straps, and light armor.',
     '+1 Production, +1 Defense', 'Light armor', 'Tannery', 'cattle', {"Production": 1.0, "Defense": 1.0}, {}),
    ('furs', 'Fibers', 'Furs', 'Furs from hunts and cold country; luxury and winter wear.',
     '+1 Culture, +1 Intelligence', '', 'Import', None, {"Culture": 1.0, "Intelligence": 1.0}, {}),
    ('dyes', 'Crafts', 'Dyes', 'Dyes and fixatives for cloth and banners.',
     '+1 Culture, +1 Production', 'Dyed cloth', "Dyer's workshop", 'dye source (woad etc.)', {"Culture": 1.0, "Production": 1.0}, {}),
    ('silk', 'Fibers', 'Silk', 'Silk thread and fabric; expensive luxury import.',
     '+2 Culture', '', 'Import', None, {"Culture": 2.0}, {}),
    ('soft-metals', 'Mines', 'Soft metals (Cu, Sn, Pb)', 'Copper, tin, and lead for tools, alloys, and simple weapons.',
     '+1 Production', 'Simple weapons (poor quality)', 'Mine - soft metals', 'soft metals deposit', {"Production": 1.0}, {}),
    ('iron', 'Mines', 'Iron', 'Iron ore and blooms; foundation of tools and military arms.',
     '+1 Production, +1 Defense', 'Military weapons, normal quality, Blacksmith Workshop', 'Mine - Iron', 'iron deposit', {"Production": 1.0, "Defense": 1.0}, {}),
    ('silver', 'Mines', 'Silver', 'Silver for coinage, jewelry, and treasury.',
     '+1 Economy, +1 Culture', 'Mint, jeweler', 'Mine - Silver', 'silver deposit', {"Economy": 1.0, "Culture": 1.0}, {}),
    ('gold', 'Mines', 'Gold', 'Precious metal for coins, regalia, and prestige.',
     '+1 Economy, +1 Culture', 'Mint, jeweler', 'Mine - Gold', 'gold deposit', {"Economy": 1.0, "Culture": 1.0}, {}),
    ('gemstones', 'Mines', 'Gemstones', 'Precious and ornamental stones.',
     '+1 Culture, +1 Magic', 'Jeweler', 'Mine - precious gems (luxury)', None, {"Culture": 1.0, "Magic": 1.0}, {}),
    ('building-stone', 'Mines', 'Building stone', 'Common stone for walls, houses, and roads.',
     '+1 Production, +1 Defense', 'Larger buildings, walls', 'Quarry - common stone', 'stone deposit', {"Production": 1.0, "Defense": 1.0}, {}),
    ('granite', 'Mines', 'Granite', 'Hard stone for towers and fortress walls.',
     '+2 Defense', 'Keeps, towers, defensive walls', 'Quarry - Granite', 'granite deposit', {"Defense": 2.0}, {}),
    ('clay', 'Mines', 'Clay', 'Clay from pits; raw material for bricks and pottery.',
     '+1 Production, +1 Economy', 'Pottery workshop, brickyard', 'Clay pit', 'clay deposit', {"Production": 1.0, "Economy": 1.0}, {}),
    ('bricks', 'Crafts', 'Bricks', 'Fired brick for durable masonry.',
     '+1 Production, +1 Defense', 'Masonry buildings and walls', 'Brickyard', 'clay', {"Production": 1.0, "Defense": 1.0}, {}),
    ('ceramics', 'Crafts', 'Ceramics', 'Pots, tiles, and glazed wares.',
     '+1 Culture, +1 Loyalty', '', "Potter's workshop", 'clay', {"Culture": 1.0, "Loyalty": 1.0}, {}),
    ('shipbuilding-wood', 'Wood', 'Shipbuilding timber', 'Timber for hulls, masts, and shipyards.',
     '+1 Production, +1 Defense', 'Shipyards', 'Sawmill - Shipbuilding wood', 'shipbuilding wood', {"Production": 1.0, "Defense": 1.0}, {}),
    ('ironwood', 'Wood', 'Ironwood', 'Very hard, rare timber for superior crafts and structures.',
     '+1 Production, +1 Defense', 'Weapons (good quality)', 'Sawmill - Ironwood', 'ironwood forest', {"Production": 1.0, "Defense": 1.0}, {}),
    ('elven-alder', 'Wood', 'Elven alder', 'Rare bow- and luxury-grade wood from deep forests.',
     '+1 Culture, +1 Magic', '', 'Sawmill - Elven alder', 'elven alder', {"Culture": 1.0, "Magic": 1.0}, {}),
    ('herbs-roots', 'Wood', 'Herbs & roots', 'Medicinal and alchemical plants from woods and gardens.',
     '+1 Science, +1 Magic', '', "Herbalist / Alchemist's Workshop", 'dense forest', {"Science": 1.0, "Magic": 1.0}, {}),
    ('paper', 'Crafts', 'Paper', 'Paper for records, letters, and scriptoria.',
     '+1 Science, +1 Intelligence', 'Scriptorium', 'Paper mill', None, {"Science": 1.0, "Intelligence": 1.0}, {}),
    ('glass', 'Crafts', 'Glass', 'Glass for windows, vessels, and instruments.',
     '+1 Culture, +1 Science', '', 'Glassworks', 'sand/gravel pit', {"Culture": 1.0, "Science": 1.0}, {}),
    ('spices', 'Exotic', 'Spices', 'Pepper, ginger, cinnamon, and similar distant imports.',
     '+2 Economy', '', 'Import', None, {"Economy": 2.0}, {}),
    ('sugar', 'Exotic', 'Sugar', 'Cane sugar; costly sweetener and preserves.',
     '+1 Culture, +1 Loyalty', '', 'Import', None, {"Culture": 1.0, "Loyalty": 1.0}, {}),
    ('amber', 'Crafts', 'Amber', 'Fossil resin from shores; jewelry and charms.',
     '+1 Culture, +1 Magic', 'Jeweler', 'Amber gatherer', 'amber coast', {"Culture": 1.0, "Magic": 1.0}, {}),
    ('ivory', 'Exotic', 'Ivory & walrus bone', 'Carving stock for luxury goods and ornament.',
     '+1 Culture, +1 Production', '', 'Import', None, {"Culture": 1.0, "Production": 1.0}, {}),
    ('horses', 'Military', 'Horses', 'Riding stock and light cavalry mounts.',
     '+1 Defense, +1 Production', 'Light cavalry', 'Stables', 'fertile land', {"Defense": 1.0, "Production": 1.0}, {}),
    ('war-horses', 'Military', 'War horses', 'Trained chargers for medium and heavy cavalry.',
     '+2 Defense', 'Medium and heavy cavalry', 'Stables', None, {"Defense": 2.0}, {}),
    ('noble-horses', 'Military', 'Noble horses', 'Show and tourney breeds.',
     '+1 Culture, +1 Loyalty', '', 'Stables', None, {"Culture": 1.0, "Loyalty": 1.0}, {}),
    ('sulfur', 'Mines', 'Sulfur', 'Sulfur from mines; gunpowder and alchemy.',
     '+1 Production, +1% Food', 'Firearms', 'Mine - Sulfur', 'sulfur deposit', {"Production": 1.0}, {"Food": 1.0}),
    ('saltpeter', 'Crafts', 'Saltpeter', 'Saltpeter for gunpowder; often produced near towns.',
     '+1 Production, +1% Food', 'Firearms', 'Saltpeter works', 'town', {"Production": 1.0}, {"Food": 1.0}),
    ('access-arms-military', 'Military', 'Military weapons', 'Standing access to military-grade arms: swords, polearms, war bows.',
     '+2 Defense', 'Units: military weapons', 'Blacksmith Workshop', 'iron', {"Defense": 2.0}, {}),
    ('access-arms-firearms', 'Military', 'Firearms', 'Access to powder weapons: muskets and arquebuses.',
     '+2 Defense', 'Units: powder weapons', 'Gunsmith workshop', 'sulfur, saltpeter', {"Defense": 2.0}, {}),
    ('access-armor-light', 'Military', 'Light armor', 'Leather, caps, and padded protection.',
     '+2 Defense', 'Light armor', 'Tannery', 'leather', {"Defense": 2.0}, {}),
    ('access-armor-medium', 'Military', 'Medium armor', 'Mail, cuirasses, and brigandine.',
     '+2 Defense', 'Medium armor', 'Armorer (plate worker)', 'leather, iron', {"Defense": 2.0}, {}),
    ('access-armor-heavy', 'Military', 'Heavy armor', 'Plate and partial plate harness.',
     '+2 Defense', 'Heavy armor', 'Armorer (plate worker)', 'leather, iron', {"Defense": 2.0}, {}),
    ('obsidian', 'Mines', 'Obsidian', 'Volcanic glass; sharp tools and ritual blades.',
     '+1 Magic, +1 Production', '', 'Quarry - Obsidian', 'obsidian deposit', {"Magic": 1.0, "Production": 1.0}, {}),
    ('tarnit', 'Mines', 'Tarnit', 'Rare stone prized for the strongest fortifications.',
     '+1 Magic, +1 Defense', 'Top-tier fortifications', 'Quarry - Tarnit', 'tarnit deposit', {"Magic": 1.0, "Defense": 1.0}, {}),
    ('dagoferryt', 'Mines', 'Dagoferryt', 'Rare metal for the finest weapons and armor.',
     '+2 Defense', 'Good-quality weapons', 'Mine - Dagoferryt', 'dagoferryt deposit', {"Defense": 2.0}, {}),
    ('elf-forest-goods', 'Exotic', 'Elven forest crafts', 'Finished luxury goods from distant elven woods.',
     '+1 Culture, +1 Magic', '', 'Import', None, {"Culture": 1.0, "Magic": 1.0}, {}),
]

SECTION_MAP = {
    "Food": "TradeGoodSection.Food",
    "Fibers": "TradeGoodSection.Fibers",
    "Mines": "TradeGoodSection.Mines",
    "Wood": "TradeGoodSection.Wood",
    "Crafts": "TradeGoodSection.Crafts",
    "Exotic": "TradeGoodSection.Exotic",
    "Military": "TradeGoodSection.Military",
}

def main():
    goods = load_goods_from_xlsx() if XLSX.is_file() else _GOODS_FALLBACK
    lines = [HEADER]
    for row in goods:
        key, sec, name, desc, bonus, unlocks, bld, req, add, pct = row
        sec_cs = SECTION_MAP[sec]
        lines.append(f'            {{ var g = G("{key}", {sec_cs}, {esc(name)}, {esc(desc)},')
        lines.append(f'                {esc(bonus)}, {esc(unlocks)}, {esc(bld)}, {esc(req)});')
        for k, v in add.items():
            lines.append(f'            g.BonusAdditive[Ppb.{k}] = {v}m;')
        for k, v in pct.items():
            lines.append(f'            g.BonusPercent[Ppb.{k}] = {v}m;')
        lines.append('            }')
    lines.append(FOOTER)
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT} ({len(goods)} goods) from {(XLSX.name if XLSX.is_file() else 'fallback')}")

if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
Export all BuildingTemplateSeeder entries into Katalog_towarow_handlowych.xlsx → Buildings.

Merges with existing sheet rows:
- App buildings (by App template name / Name) get In app catalog=Y and current costs/PPB/notes from seeder
- Excel-only rows (planned / aliases not in seeder) are kept with In app catalog=N
- Prefer Excel "Name (trade goods)" when it already matched an app template
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs/barony/Katalog_towarow_handlowych.xlsx"
SEEDER = ROOT / "DagoniteEmpire/Service/BuildingTemplateSeeder.cs"

HEADERS = [
    "Name (trade goods)",
    "App template name",
    "In app catalog",
    "Lordship",
    "Kind",
    "Production cost",
    "Gold cost",
    "Food",
    "Economy",
    "Production",
    "Loyalty",
    "Stability",
    "Law",
    "Corruption",
    "Science",
    "Magic",
    "Culture",
    "Intelligence",
    "Defense",
    "Treasury",
    "Notes",
]

PPB_ARGS = [
    "food",
    "economy",
    "production",
    "loyalty",
    "stability",
    "law",
    "corruption",
    "science",
    "magic",
    "culture",
    "intelligence",
    "defense",
    "treasury",
]

PPB_HEADERS = [
    "Food",
    "Economy",
    "Production",
    "Loyalty",
    "Stability",
    "Law",
    "Corruption",
    "Science",
    "Magic",
    "Culture",
    "Intelligence",
    "Defense",
    "Treasury",
]

# Excel App template name / Name aliases → seeder Name
ALIASES = {
    "armorer": "Armorer",
    "armorer (plate worker)": "Armorer",
    "armorers guild": "Plate Workshop",
    "cheesemaker": "Cheese dairy",
    "cheese dairy": "Cheese dairy",
    "farm – cotton": "Cotton farm",
    "farm - cotton": "Cotton farm",
    "farm – flax": "Flax farm",
    "farm - flax": "Flax farm",
    "farm - woad": "Farm (Dye plant)",
    "farm - madder": "Farm (Dye plant)",
    "farm - weld": "Farm (Dye plant)",
    "woad farm": "Farm (Dye plant)",
    "madder farm": "Farm (Dye plant)",
    "weld farm": "Farm (Dye plant)",
    "dye garden": "Farm (Dye plant)",
    "farm (dye plant)": "Farm (Dye plant)",
    "farm - dye plant": "Farm (Dye plant)",
    "farm – dye plant": "Farm (Dye plant)",
    "herbalist": "Herbalist / Alchemist's Workshop",
    "meat saltery": "Meat saltery (salt + fish/meat)",
    "candlemaker": "Candlemaker",
    "horse stud (militart)": "Horse Stud (military)",
}


def unescape_cs(s: str) -> str:
    return (
        s.replace("\\n", "\n")
        .replace('\\"', '"')
        .replace("\\\\", "\\")
    )


def parse_seeder() -> list[dict]:
    text = SEEDER.read_text(encoding="utf-8")
    blocks: list[tuple[str, str]] = []
    for m in re.finditer(r'Add\(\s*"((?:[^"\\]|\\.)*)"\s*,', text):
        name = m.group(1).replace('\\"', '"')
        start = m.start()
        depth = 0
        i = start
        while i < len(text):
            ch = text[i]
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
                if depth == 0 and i + 1 < len(text) and text[i + 1] == ";":
                    blocks.append((name, text[start : i + 2]))
                    break
            i += 1

    rows: list[dict] = []
    for name, block in blocks:
        # Add("Name", lordship, B|I, prod, gold, "desc", Fx(...), terrain?);
        head = re.match(
            r'Add\(\s*"(?:[^"\\]|\\.)*"\s*,\s*(\d+)\s*,\s*([BI])\s*,\s*(\d+)\s*,\s*(\d+)\s*,',
            block,
        )
        if not head:
            # additive-only form: Add("Palace", 2, B, additive: Fx(...));
            head2 = re.match(
                r'Add\(\s*"(?:[^"\\]|\\.)*"\s*,\s*(\d+)\s*,\s*([BI])\s*,\s*additive:\s*Fx\(([^)]*)\)\s*\);',
                block,
            )
            if not head2:
                raise RuntimeError(f"Cannot parse Add block for {name}:\n{block[:200]}")
            lordship = int(head2.group(1))
            kind = "Building" if head2.group(2) == "B" else "Improvement"
            prod = 0
            gold = 0
            desc = ""
            fx_body = head2.group(3)
        else:
            lordship = int(head.group(1))
            kind = "Building" if head.group(2) == "B" else "Improvement"
            prod = int(head.group(3))
            gold = int(head.group(4))
            desc_m = re.search(r',\s*"((?:[^"\\]|\\.)*)"\s*(?:,|\))', block[head.end() - 1 :])
            # Better: first string after costs
            desc_m = re.search(
                r'Add\(\s*"(?:[^"\\]|\\.)*"\s*,\s*\d+\s*,\s*[BI]\s*,\s*\d+\s*,\s*\d+\s*,\s*"((?:[^"\\]|\\.)*)"',
                block,
            )
            desc = unescape_cs(desc_m.group(1)) if desc_m else ""
            fx_m = re.search(r"Fx\(([^)]*)\)", block)
            fx_body = fx_m.group(1) if fx_m else ""

        ppb = {h: None for h in PPB_HEADERS}
        for arg, header in zip(PPB_ARGS, PPB_HEADERS):
            m = re.search(rf"\b{arg}\s*:\s*(-?\d+(?:\.\d+)?m?)", fx_body)
            if not m:
                continue
            raw = m.group(1).rstrip("m")
            val = float(raw)
            ppb[header] = int(val) if val == int(val) else val

        rows.append(
            {
                "Name (trade goods)": name,
                "App template name": name,
                "In app catalog": "Y",
                "Lordship": lordship,
                "Kind": kind,
                "Production cost": prod,
                "Gold cost": gold,
                **ppb,
                "Notes": desc,
            }
        )
    return rows


def load_existing_excel() -> list[dict]:
    if not XLSX.exists():
        return []
    wb = openpyxl.load_workbook(XLSX)
    if "Buildings" not in wb.sheetnames:
        return []
    ws = wb["Buildings"]
    hdr = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            hdr[str(v).strip()] = c
    rows = []
    for r in range(2, ws.max_row + 1):
        trade = ws.cell(r, hdr.get("Name (trade goods)", 1)).value
        app = ws.cell(r, hdr.get("App template name", 2)).value
        if not trade and not app:
            continue
        row = {h: None for h in HEADERS}
        for h, c in hdr.items():
            if h in row:
                row[h] = ws.cell(r, c).value
        if row.get("Name (trade goods)"):
            row["Name (trade goods)"] = str(row["Name (trade goods)"]).strip()
        if row.get("App template name"):
            row["App template name"] = str(row["App template name"]).strip()
        rows.append(row)
    return rows


def norm(s: str | None) -> str:
    if not s:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower().replace("–", "-").replace("—", "-"))


def resolve_app_key(row: dict, seeder_by_name: dict[str, dict]) -> str | None:
    candidates = []
    for key in ("App template name", "Name (trade goods)"):
        v = row.get(key)
        if v:
            candidates.append(str(v).strip())
    for c in candidates:
        if c in seeder_by_name:
            return c
        alias = ALIASES.get(norm(c))
        if alias and alias in seeder_by_name:
            return alias
        # case-insensitive
        for name in seeder_by_name:
            if name.lower() == c.lower():
                return name
    return None


def excel_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).strip().replace(",", "."))
    except ValueError:
        return v


def merge(seeder_rows: list[dict], excel_rows: list[dict]) -> list[dict]:
    seeder_by_name = {r["App template name"]: r for r in seeder_rows}
    matched_seeder: set[str] = set()
    out: list[dict] = []

    # 1) Keep excel order for matched / excel-only rows
    for er in excel_rows:
        key = resolve_app_key(er, seeder_by_name)
        if key is None:
            # Planned / not in app — keep as-is
            row = {h: er.get(h) for h in HEADERS}
            row["In app catalog"] = "N"
            if not row.get("App template name") and row.get("Name (trade goods)"):
                row["App template name"] = row["Name (trade goods)"]
            if not row.get("Name (trade goods)") and row.get("App template name"):
                row["Name (trade goods)"] = row["App template name"]
            out.append(row)
            continue

        sr = seeder_by_name[key]
        matched_seeder.add(key)
        row = dict(sr)
        # Preserve trade-goods display name from Excel when present
        if er.get("Name (trade goods)"):
            row["Name (trade goods)"] = str(er["Name (trade goods)"]).strip()
        row["App template name"] = key
        row["In app catalog"] = "Y"
        out.append(row)

    # 2) Append seeder buildings missing from Excel
    for name, sr in seeder_by_name.items():
        if name in matched_seeder:
            continue
        out.append(dict(sr))

    # Stable sort: Kind (Building then Improvement), then name
    kind_order = {"Building": 0, "Improvement": 1}

    def sort_key(r: dict):
        return (
            kind_order.get(str(r.get("Kind") or ""), 9),
            norm(r.get("App template name") or r.get("Name (trade goods)")),
        )

    out.sort(key=sort_key)
    return out


def write_buildings(rows: list[dict]) -> None:
    wb = openpyxl.load_workbook(XLSX)
    if "Buildings" in wb.sheetnames:
        del wb["Buildings"]
    ws = wb.create_sheet("Buildings")
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)

    for r_i, row in enumerate(rows, 2):
        for c, h in enumerate(HEADERS, 1):
            v = row.get(h)
            if h in PPB_HEADERS or h in ("Production cost", "Gold cost", "Lordship"):
                v = excel_num(v)
            ws.cell(r_i, c, v)

    ws.auto_filter.ref = f"A1:U{len(rows) + 1}"
    ws.freeze_panes = "A2"
    widths = {
        "A": 28,
        "B": 32,
        "C": 12,
        "D": 10,
        "E": 12,
        "F": 14,
        "G": 10,
        "U": 60,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for col in "HIJKLMNOPQRST":
        ws.column_dimensions[col].width = 11

    wb.save(XLSX)


def main() -> None:
    seeder_rows = parse_seeder()
    excel_rows = load_existing_excel()
    merged = merge(seeder_rows, excel_rows)
    write_buildings(merged)
    in_app = sum(1 for r in merged if str(r.get("In app catalog")).upper() == "Y")
    planned = len(merged) - in_app
    print(f"Wrote {XLSX}")
    print(f"  Seeder buildings: {len(seeder_rows)}")
    print(f"  Rows in Buildings sheet: {len(merged)} (in app: {in_app}, planned/excel-only: {planned})")


if __name__ == "__main__":
    main()
